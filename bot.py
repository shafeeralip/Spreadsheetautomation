import openpyxl as xl
from openpyxl.chart import Reference, BarChart


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    corrected_cell_heading = sheet.cell(1, 4)
    corrected_cell_heading.value = 'prices'

    for row in range(2, sheet.max_row + 1):
        first_price = sheet.cell(row, 3)
        corrected_price = first_price.value * 10
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'f5')
    wb.save(filename)


process_workbook('Book1.xlsx')








